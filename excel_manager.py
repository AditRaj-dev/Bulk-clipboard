import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from config import EXCEL_FILE, CODE_COLUMN, FONT_NAME


def _cursor_file(sheet):
    return f"cursor_{sheet.replace(' ', '_').lower()}.txt"


def get_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.remove(wb.active)
        wb.save(EXCEL_FILE)
    return load_workbook(EXCEL_FILE)


def get_sheet(sheet_name):
    wb = get_workbook()
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        wb.save(EXCEL_FILE)
    return wb, wb[sheet_name]


def append_text(sheet, text):
    wb, ws = get_sheet(sheet)
    row = 1
    while ws.cell(row=row, column=CODE_COLUMN).value:
        row += 1
    cell = ws.cell(row=row, column=CODE_COLUMN)
    cell.value = text
    cell.alignment = Alignment(wrap_text=True)
    cell.font = Font(name=FONT_NAME)
    wb.save(EXCEL_FILE)
    return row


def read_row(sheet, row):
    _, ws = get_sheet(sheet)
    return ws.cell(row=row, column=CODE_COLUMN).value


def get_cursor(sheet, default):
    f = _cursor_file(sheet)
    if not os.path.exists(f):
        set_cursor(sheet, default)
        return default
    return int(open(f).read())


def set_cursor(sheet, row):
    with open(_cursor_file(sheet), "w") as f:
        f.write(str(row))


def clear_cursor(sheet):
    f = _cursor_file(sheet)
    if os.path.exists(f):
        os.remove(f)

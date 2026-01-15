import os
from openpyxl import Workbook, load_workbook
from config import EXCEL_FILE, SHEETS


def ensure_sheets():
    created = []
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        # remove default sheet
        if wb.sheetnames:
            wb.remove(wb.active)
        for name in SHEETS:
            wb.create_sheet(name)
            created.append(name)
        wb.save(EXCEL_FILE)
        print(f"Created {EXCEL_FILE} with sheets: {', '.join(created)}")
        return

    wb = load_workbook(EXCEL_FILE)
    for name in SHEETS:
        if name not in wb.sheetnames:
            wb.create_sheet(name)
            created.append(name)
    # remove default 'Sheet' if it exists and is empty and not in SHEETS
    default_name = 'Sheet'
    if default_name in wb.sheetnames and default_name not in SHEETS:
        ws = wb[default_name]
        if ws.max_row == 1 and ws.max_column == 1 and ws['A1'].value is None:
            wb.remove(ws)
    if created:
        wb.save(EXCEL_FILE)
        print(f"Added sheets to {EXCEL_FILE}: {', '.join(created)}")
    else:
        print(f"No sheets added. {EXCEL_FILE} already contains all configured sheets.")


if __name__ == '__main__':
    ensure_sheets()

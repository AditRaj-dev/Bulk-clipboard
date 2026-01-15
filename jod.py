import os
import time
import threading
import pyperclip
import keyboard
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
import tkinter as tk
from tkinter import ttk, messagebox

# ================= CONFIG =================
EXCEL_FILE = "java_snippets.xlsx"
SHEET_NAME = "Code"
CODE_COLUMN = 1
FONT_NAME = "Consolas"
CURSOR_FILE = "excel_cursor.txt"
AHK_HOTKEY = "ctrl+alt+1"
# ==========================================

# ================= DARK THEME =================
BG = "#1e1e1e"
PANEL = "#252526"
BTN = "#3c3c3c"
FG = "#d4d4d4"
ENTRY_BG = "#2d2d2d"
SCROLL_BG = "#3a3a3a"
SCROLL_THUMB = "#5a5a5a"
# =============================================

loading_active = False
waiting_for_paste = False
watching = False


# ---------- Cursor Helpers ----------
def set_cursor(row: int):
    with open(CURSOR_FILE, "w") as f:
        f.write(str(row))
    current_row_var.set(str(row))


def clear_cursor():
    if os.path.exists(CURSOR_FILE):
        os.remove(CURSOR_FILE)
    current_row_var.set("-")


def get_cursor(default_row: int):
    if not os.path.exists(CURSOR_FILE):
        set_cursor(default_row)
        return default_row
    row = int(open(CURSOR_FILE).read())
    current_row_var.set(str(row))
    return row


# ---------- Excel Utilities ----------
def get_sheet():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    return wb, wb[SHEET_NAME]


def save_to_excel(text):
    wb, ws = get_sheet()
    row = 1
    while ws.cell(row=row, column=CODE_COLUMN).value:
        row += 1
    cell = ws.cell(row=row, column=CODE_COLUMN)
    cell.value = text
    cell.alignment = Alignment(wrap_text=True)
    cell.font = Font(name=FONT_NAME)
    wb.save(EXCEL_FILE)
    return row


# ---------- Loading Logic ----------
def load_current_row(start_row):
    global waiting_for_paste

    current_row = get_cursor(start_row)
    wb, ws = get_sheet()
    value = ws.cell(row=current_row, column=CODE_COLUMN).value

    if not value:
        log_message(f"⚠ No snippet at row {current_row}")
        stop_loading()
        return

    pyperclip.copy(value)
    waiting_for_paste = True
    log_message(f"📋 Loaded row {current_row} → clipboard (waiting for paste)")


def start_loading():
    global loading_active

    if loading_active:
        return

    if not os.path.exists(CURSOR_FILE):
        set_start_row()

    loading_active = True
    log_message("▶ Continuous loading started")
    load_current_row(int(current_row_var.get()))


def stop_loading():
    global loading_active, waiting_for_paste

    loading_active = False
    waiting_for_paste = False
    clear_cursor()
    log_message("⏸ Loading stopped and cursor reset")


def on_paste_detected():
    global waiting_for_paste

    if not loading_active or not waiting_for_paste:
        return

    current_row = int(current_row_var.get())
    next_row = current_row + 1
    set_cursor(next_row)
    waiting_for_paste = False
    log_message(f"➡️ Paste detected → moved to row {next_row}")

    if loading_active:
        load_current_row(next_row)


# ---------- Start Row Confirmation ----------
def set_start_row():
    try:
        row = int(start_row_var.get())
        if row < 1:
            raise ValueError
    except ValueError:
        messagebox.showerror("Invalid Row", "Start row must be a positive integer.")
        return

    clear_cursor()
    set_cursor(row)
    log_message(f"✅ Start row confirmed → cursor set to row {row}")


# ---------- Clipboard Capture ----------
def watch_clipboard(log):
    global watching
    try:
        last_text = pyperclip.paste()
    except Exception:
        last_text = None

    log("Clipboard capture started")

    while watching:
        try:
            text = pyperclip.paste()
        except Exception:
            text = None

        if text and text != last_text:
            row = save_to_excel(text)
            log(f"Saved clipboard → Excel row {row}")
            last_text = text

        time.sleep(0.5)

    log("Clipboard capture stopped")


def start_capture():
    global watching
    if watching:
        return
    watching = True
    threading.Thread(target=watch_clipboard, args=(log_message,), daemon=True).start()
    log_message("▶ Capture started")


def stop_capture():
    global watching
    watching = False
    log_message("⏹ Capture stopped")


# ---------- Misc ----------
def reset_excel_confirm():
    if messagebox.askyesno("Reset Excel", "This will DELETE all snippets.\nContinue?"):
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
        clear_cursor()
        log_message("🧹 Excel reset successfully")


def clear_logs():
    log_text.delete("1.0", tk.END)
    log_message("Logs cleared")


def trigger_ahk_hotkey():
    keyboard.press_and_release(AHK_HOTKEY)
    log_message(f"⌨️ Triggered AHK hotkey ({AHK_HOTKEY})")


def log_message(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    log_text.insert(tk.END, f"[{ts}] {msg}\n")
    log_text.see(tk.END)


# ---------- GUI ----------
root = tk.Tk()
root.title("Excel ⇄ Clipboard Code Manager")
root.geometry("1060x650")
root.configure(bg=BG)

keyboard.add_hotkey("ctrl+v", on_paste_detected)

style = ttk.Style()
style.theme_use("default")
style.configure(
    "Vertical.TScrollbar",
    background=SCROLL_BG,
    troughcolor=PANEL,
    arrowcolor=FG,
    relief="flat"
)

control_frame = tk.Frame(root, bg=PANEL)
control_frame.pack(fill=tk.X, pady=6)

tk.Label(control_frame, text="Start Row:", bg=PANEL, fg=FG).pack(side=tk.LEFT, padx=6)
start_row_var = tk.StringVar(value="1")
tk.Entry(
    control_frame,
    textvariable=start_row_var,
    width=6,
    bg=ENTRY_BG,
    fg=FG,
    insertbackground=FG,
    relief=tk.FLAT
).pack(side=tk.LEFT)

tk.Button(
    control_frame,
    text="Set",
    command=set_start_row,
    bg=BTN,
    fg=FG,
    relief=tk.FLAT
).pack(side=tk.LEFT, padx=6)

tk.Label(control_frame, text="Current Row:", bg=PANEL, fg=FG).pack(side=tk.LEFT, padx=20)
current_row_var = tk.StringVar(value="-")
tk.Entry(
    control_frame,
    textvariable=current_row_var,
    width=6,
    state="readonly",
    readonlybackground=ENTRY_BG,
    fg=FG,
    relief=tk.FLAT
).pack(side=tk.LEFT)

btn_row_1 = tk.Frame(root, bg=PANEL)
btn_row_1.pack(pady=6)

btn_row_2 = tk.Frame(root, bg=PANEL)
btn_row_2.pack(pady=6)

def dark_btn(parent, text, cmd, w=18):
    return tk.Button(
        parent,
        text=text,
        width=w,
        command=cmd,
        bg=BTN,
        fg=FG,
        activebackground=PANEL,
        activeforeground=FG,
        relief=tk.FLAT
    )

dark_btn(btn_row_1, "▶ Start Capture", start_capture).pack(side=tk.LEFT, padx=5)
dark_btn(btn_row_1, "⏹ Stop Capture", stop_capture).pack(side=tk.LEFT, padx=5)
dark_btn(btn_row_1, "▶ Start Loading", start_loading, 22).pack(side=tk.LEFT, padx=5)
dark_btn(btn_row_1, "⏸ Stop Loading", stop_loading).pack(side=tk.LEFT, padx=5)

dark_btn(btn_row_2, "⌨️ Trigger AHK", trigger_ahk_hotkey).pack(side=tk.LEFT, padx=5)
dark_btn(btn_row_2, "🧹 Reset Excel Sheet", reset_excel_confirm, 22).pack(side=tk.LEFT, padx=5)
dark_btn(btn_row_2, "🧾 Clear Logs", clear_logs).pack(side=tk.LEFT, padx=5)

log_frame = tk.Frame(root, bg=BG)
log_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

log_text = tk.Text(
    log_frame,
    bg=BG,
    fg=FG,
    insertbackground=FG,
    font=("Consolas", 10),
    relief=tk.FLAT,
    wrap=tk.WORD
)
log_text.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

log_scroll = ttk.Scrollbar(
    log_frame,
    orient="vertical",
    command=log_text.yview,
    style="Vertical.TScrollbar"
)
log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
log_text.configure(yscrollcommand=log_scroll.set)

log_message("Ready. Cursor reset + Start Row confirmation fixed.")

root.mainloop()
